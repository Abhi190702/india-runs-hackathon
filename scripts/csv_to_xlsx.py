# -*- coding: utf-8 -*-
"""Convert submission.csv -> submission.xlsx (same 4 columns, correct types)."""
import csv
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

src = sys.argv[1] if len(sys.argv) > 1 else "submission.csv"
out = sys.argv[2] if len(sys.argv) > 2 else "submission.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "ranking"

with open(src, encoding="utf-8") as f:
    reader = csv.reader(f)
    header = next(reader)
    ws.append(header)  # candidate_id, rank, score, reasoning
    n = 0
    for row in reader:
        cid, rank, score, reasoning = row[0], int(row[1]), float(row[2]), row[3]
        ws.append([cid, rank, score, reasoning])
        n += 1

# --- presentation: readable widths, bold header, frozen header row ----------
ws.column_dimensions["A"].width = 16   # candidate_id
ws.column_dimensions["B"].width = 7    # rank
ws.column_dimensions["C"].width = 11   # score
ws.column_dimensions["D"].width = 120  # reasoning

for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")

# left-align the reasoning column so it reads cleanly
for r in range(2, n + 2):
    ws.cell(r, 4).alignment = Alignment(horizontal="left", vertical="center")

ws.freeze_panes = "A2"  # keep header visible while scrolling

wb.save(out)
print(f"Wrote {n} rows to {out} (formatted widths + bold header)")
